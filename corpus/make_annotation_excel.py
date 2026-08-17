"""
make_annotation_excel.py - Generateur de fichier Excel ergonomique pour annotation
====================================================================================
Convertit un fichier CSV d'annotation en fichier Excel avec :
  - Menus deroulants pour edu_emotion et metacog_label
  - Colonnes figees apres 'utterance' (plus de scroll horizontal)
  - Lignes overlap en fond jaune
  - Onglet 'Guide rapide' inclus

Usage:
    pip install openpyxl
    python make_annotation_excel.py to_annotate_zinedine.csv
    python make_annotation_excel.py to_annotate_sara.csv

Auteurs : Zinedine Hamadi & Sara Hadidi
Memoire : M1 Sciences du Langage et Informatique - Sorbonne Universite 2025/2026
"""

import sys
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

INPUT_FILE  = sys.argv[1] if len(sys.argv) > 1 else "to_annotate_zinedine.csv"
OUTPUT_FILE = INPUT_FILE.replace(".csv", "_annotation.xlsx")

EMO_VALUES  = "stress,frustration,engagement,confusion,satisfaction,neutre,autre"
META_VALUES = "planification,monitoring,auto-eval,aucun"

C_HEADER   = "1F3864"
C_READONLY = "E8EDF4"
C_ANNOT    = "E8F5E9"
C_OVERLAP  = "FFF8E1"

thin   = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

print(f"Chargement de {INPUT_FILE}...")
df = pd.read_csv(INPUT_FILE, encoding="utf-8")
print(f"  -> {len(df)} lignes")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Annotation"

headers    = list(df.columns)
annot_cols = ["edu_emotion", "metacog_label", "notes"]

# En-tetes
for col_idx, col_name in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font      = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    cell.fill      = PatternFill("solid", fgColor=C_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = border
ws.row_dimensions[1].height = 30

# Donnees
for row_idx, row in df.iterrows():
    excel_row  = row_idx + 2
    is_overlap = row_idx < 50

    for col_idx, col_name in enumerate(headers, start=1):
        value = row[col_name] if pd.notna(row[col_name]) else ""
        cell  = ws.cell(row=excel_row, column=col_idx, value=value)
        cell.border = border
        cell.font   = Font(size=10, name="Arial")

        if col_name in annot_cols:
            fill_color = "D4EDDA" if is_overlap else C_ANNOT
        else:
            fill_color = "FFF3CD" if is_overlap else C_READONLY

        cell.fill      = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.row_dimensions[excel_row].height = 45

# Largeurs
col_widths = {
    "source": 18, "conv_id": 16, "context": 45, "utterance": 45,
    "emotion": 14, "matched_kw": 22, "annotator": 16,
    "edu_emotion": 16, "metacog_label": 16, "notes": 30,
}
for col_idx, col_name in enumerate(headers, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 15)

# Menus deroulants
emo_col  = headers.index("edu_emotion")  + 1
meta_col = headers.index("metacog_label") + 1

dv_emo = DataValidation(
    type="list", formula1=f'"{EMO_VALUES}"', allow_blank=True, showDropDown=False,
)
dv_meta = DataValidation(
    type="list", formula1=f'"{META_VALUES}"', allow_blank=True, showDropDown=False,
)
dv_emo.sqref  = f"{get_column_letter(emo_col)}2:{get_column_letter(emo_col)}{len(df)+1}"
dv_meta.sqref = f"{get_column_letter(meta_col)}2:{get_column_letter(meta_col)}{len(df)+1}"
ws.add_data_validation(dv_emo)
ws.add_data_validation(dv_meta)

# Colonnes figees apres utterance
utt_col = headers.index("utterance") + 2
ws.freeze_panes = f"{get_column_letter(utt_col)}2"

# Onglet guide rapide
ws2 = wb.create_sheet(title="Guide rapide")
guide = [
    ["GUIDE D'ANNOTATION RAPIDE", ""],
    ["", ""],
    ["edu_emotion",    "Definition"],
    ["stress",         "Tension, pression face a une tache ou echeance"],
    ["frustration",    "Blocage repete, echec sans progression visible"],
    ["engagement",     "Interet actif, curiosite, enthousiasme pour apprendre"],
    ["confusion",      "Incomprehension d'un concept, incertitude cognitive"],
    ["satisfaction",   "Reussite, comprehension atteinte, soulagement"],
    ["neutre",         "Aucune emotion educative detectable"],
    ["autre",          "Emotion presente mais hors des 5 categories educatives"],
    ["", ""],
    ["metacog_label",  "Definition"],
    ["planification",  "L'apprenant definit une strategie avant d'agir"],
    ["monitoring",     "L'apprenant surveille sa comprehension en cours"],
    ["auto-eval",      "L'apprenant evalue sa performance apres la tache"],
    ["aucun",          "Aucun marqueur metacognitif detectable"],
    ["", ""],
    ["REGLE D'OR",     "Annoter l'utterance, pas le context."],
    ["Si doute",       "Preferer neutre/aucun plutot que forcer un label."],
    ["Overlap",        "Les 50 premieres lignes (fond jaune) : ne pas consulter l'autre annotateur avant d'avoir fini."],
]
for r, (a, b) in enumerate(guide, start=1):
    c1 = ws2.cell(row=r, column=1, value=a)
    c2 = ws2.cell(row=r, column=2, value=b)
    if a in ["edu_emotion", "metacog_label", "GUIDE D'ANNOTATION RAPIDE", "REGLE D'OR"]:
        c1.font = Font(bold=True, size=11, name="Arial", color="1F3864")
    else:
        c1.font = Font(size=10, name="Arial", bold=True)
    c2.font = Font(size=10, name="Arial")
    c2.alignment = Alignment(wrap_text=True)

ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 65

wb.save(OUTPUT_FILE)
print(f"\nFichier sauvegarde : {OUTPUT_FILE}")
print(f"  - Menus deroulants sur edu_emotion et metacog_label")
print(f"  - Colonnes figees apres 'utterance'")
print(f"  - 50 premieres lignes (overlap) en fond jaune")
print(f"  - Onglet 'Guide rapide' inclus")
